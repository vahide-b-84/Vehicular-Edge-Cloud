
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
root_folder = os.path.join(BASE_DIR, "heterogeneous_results")
output_excel = os.path.join(BASE_DIR, "Final_Result_W.xlsx")

models = {
    #"Pdqn": "Pdqn",
    "dqn": "dqn",
    #"dqn1": "dqn1",
    "original_only": "original_only",
    "greedy": "greedy"
    #"dqn3": "dqn3",
    #"dqn2": "dqn2",
}

if os.path.exists(output_excel):
    os.remove(output_excel)
wb = Workbook()
wb.remove(wb.active)

def read_metrics(path, weight=False):
    """Merge Logs and Summary by episode; optional per-episode weighting."""
    logs = pd.read_excel(path, sheet_name="Logs")[['Episode', 'Avg Reward', 'Avg Delay']]
    summ = pd.read_excel(path, sheet_name="Summary")[['episode', 'Total', 'AVG_Failure']]
    df = logs.merge(summ, left_on='Episode', right_on='episode', how='inner').drop(columns=['episode'])
    if weight:
        df['Avg Reward'] *= df['Total']
        df['Avg Delay'] *= df['Total']
        df['AVG_Failure'] *= df['Total']
    return df[['Episode', 'Avg Reward', 'Avg Delay', 'AVG_Failure', 'Total']]

def write_sheet_with_charts(sheet_name, data_dict):
    """Merge per-model data and add charts (dynamic columns)."""
    ws = wb.create_sheet(sheet_name)
    merged = None
    for model, df in data_dict.items():
        df = df[['Episode', 'Avg Reward', 'Avg Delay', 'AVG_Failure']].copy()
        df.columns = ['Episode'] + [f'{c}_{model}' for c in ['Avg Reward','Avg Delay','AVG_Failure']]
        merged = df if merged is None else pd.merge(merged, df, on='Episode', how='outer')

    if merged is None or merged.empty:
        return

    for row in dataframe_to_rows(merged, index=False, header=True):
        ws.append(row)

    def add_chart(title, ytitle, base_offset, anchor):
        chart = LineChart()
        chart.title = title
        chart.y_axis.title = ytitle
        chart.x_axis.title = 'Episode'
        chart.width = 18
        chart.height = 8
        max_row = ws.max_row
        max_col = ws.max_column
        num_models = (max_col - 1) // 3
        for i in range(num_models):
            c = 1 + base_offset + i * 3  # 2->Reward, 3->Delay, 4->Failure
            chart.add_data(Reference(ws, min_col=c, max_col=c, min_row=1, max_row=max_row), titles_from_data=True)
        cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
        chart.set_categories(cats)
        ws.add_chart(chart, anchor)

    add_chart("Avg Reward", "Reward", base_offset=1, anchor="L2")
    add_chart("Avg Delay", "Delay", base_offset=2, anchor="L20")
    add_chart("AVG_Failure", "Failure", base_offset=3, anchor="L38")

# 1) Global_model (unweighted)
global_data = {}
for model in models:
    file_path = os.path.join(root_folder, model, f"Global_Permutation_3_{model}.xlsx")
    if os.path.exists(file_path):
        df = read_metrics(file_path, weight=False)
        global_data[model] = df
write_sheet_with_charts("Global_model", global_data)

# 2) RSU sheets (weighted by per-episode Total)
rsu_ids = set()
for model in models:
    folder = os.path.join(root_folder, model)
    if not os.path.exists(folder):
        continue
    for file in os.listdir(folder):
        if file.startswith("RSU_") and file.endswith(f"{model}.xlsx"):
            rsu_id = file.split("_Permutation")[0]
            rsu_ids.add(rsu_id)
rsu_ids = sorted(rsu_ids)

weighted_parts = {m: [] for m in models}     # numerators
weight_denoms = {m: [] for m in models}      # denominators

for rsu_id in rsu_ids:
    rsu_data = {}
    for model in models:
        path = os.path.join(root_folder, model, f"{rsu_id}_Permutation_3_{model}.xlsx")
        if os.path.exists(path):
            df = read_metrics(path, weight=False)  # has Total
            df_w = df.copy()
            df_w['Avg Reward'] *= df_w['Total']
            df_w['Avg Delay'] *= df_w['Total']
            df_w['AVG_Failure'] *= df_w['Total']

            rsu_data[model] = df_w[['Episode','Avg Reward','Avg Delay','AVG_Failure']]
            weighted_parts[model].append(df_w[['Episode','Avg Reward','Avg Delay','AVG_Failure']])
            weight_denoms[model].append(df[['Episode','Total']])
    if rsu_data:
        write_sheet_with_charts(rsu_id, rsu_data)

# 3) RSU_avg (per-episode weighted average across RSUs)
rsu_avg_data = {}
for model in models:
    dfs_num = weighted_parts.get(model, [])
    dfs_den = weight_denoms.get(model, [])
    if not dfs_num or not dfs_den:
        continue
    numerators = pd.concat(dfs_num).groupby('Episode').sum(numeric_only=True).reset_index()
    denoms = pd.concat(dfs_den).groupby('Episode')['Total'].sum().reset_index()
    avg_df = numerators.merge(denoms, on='Episode', how='inner')
    avg_df = avg_df[avg_df['Total'] != 0]
    for col in ['Avg Reward','Avg Delay','AVG_Failure']:
        avg_df[col] = avg_df[col] / avg_df['Total']
    rsu_avg_data[model] = avg_df[['Episode','Avg Reward','Avg Delay','AVG_Failure']]

if rsu_avg_data:
    write_sheet_with_charts("W_RSU_avg", rsu_avg_data)

wb.save(output_excel)
print("Final_Result_W.xlsx created!")
