def save_params_and_logs(params, log_data_global, task_Assignments_info_global, rsu_logs_dict, rsu_assignments_dict):
    # Import standard libraries for file/path handling and data processing
    import os
    import pandas as pd
    import numpy as np

    # Import OpenPyXL utilities for working with Excel files, sheets, and charts
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook, load_workbook
    

    # Excel hard limits (used to prevent writing invalid-sized sheets)
    EXCEL_MAX_ROWS = 1_048_576
    EXCEL_MAX_COLS = 16_384

    def append_df_to_sheet(filename, df, sheet_name):
        # Append a DataFrame to an existing Excel sheet; create the sheet if it doesn't exist
        book = load_workbook(filename)
        if sheet_name not in book.sheetnames:
            # If the sheet does not exist, create it by writing the dataframe with headers
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay')
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.close()
        else:
            # If the sheet exists, append rows without headers
            sheet = book[sheet_name]
            for r in dataframe_to_rows(df, index=False, header=False):
                sheet.append(r)
            book.save(filename)

    def _detect_episode_col(df, candidates=("episode","Episode","EPISODE")):
        # Detect the episode column name from common candidate options
        for c in candidates:
            if c in df.columns:
                return c
        # Raise an error if no episode column is found
        raise KeyError(f"No episode column found. Tried {candidates}")

    def write_task_assignments_latest(filename: str, df, base_sheet_name="TaskAssignments", episode_col: str=None, episode_value: int=None):
        """
        # Fully replaces the contents of the 'TaskAssignments' sheet so that only the records
        # for the latest episode remain.
        # If episode_value is not provided, it uses the maximum value of the episode column in df.
        """
        # If there is no data to write, do nothing
        if df is None or len(df) == 0:
            return

        # Determine which column represents the episode
        ep_col = episode_col or _detect_episode_col(df)

        # Choose the target episode: explicit episode_value or the latest episode in the dataframe
        target_episode = episode_value if episode_value is not None else df[ep_col].max()
        df_latest = df[df[ep_col] == target_episode].copy()

        # Validate Excel column/row limits before writing
        if df_latest.shape[1] > EXCEL_MAX_COLS:
            raise ValueError(f"Too many columns for Excel: {df_latest.shape[1]} > {EXCEL_MAX_COLS}")
        if df_latest.shape[0] + 1 > EXCEL_MAX_ROWS:
            raise ValueError(f"Too many rows for Excel: {df_latest.shape[0]} data rows")

        # Try to load workbook; if missing, create a new one
        try:
            wb = load_workbook(filename)
        except FileNotFoundError:
            wb = Workbook()

        # If the target sheet already exists, remove it first to fully replace its contents
        if base_sheet_name in wb.sheetnames:
            wb.remove(wb[base_sheet_name])
            # Ensure workbook still has at least one sheet after removal
            if not wb.sheetnames:
                wb.create_sheet(title="Sheet")

        # Create the fresh sheet and write headers + data rows
        ws = wb.create_sheet(title=base_sheet_name)
        ws.append(list(df_latest.columns))
        for row in dataframe_to_rows(df_latest, index=False, header=False):
            ws.append(row)

        # Save and close workbook
        wb.save(filename)
        wb.close()


    def add_charts_to_sheet(wb, sheet_name):
        # Add (or refresh) charts for a given sheet in the workbook
        ws = wb[sheet_name]

        # Clear existing charts so updated charts don't stack
        ws._charts.clear()

        # Compute sheet bounds (used to set chart data ranges and chart placement)
        max_row = ws.max_row
        max_col = ws.max_column

        if sheet_name == 'Logs':
            # ---------------------------
            # Chart 1: Reward per Episode
            # ---------------------------
            chart_rewards = LineChart()
            chart_rewards.title = "Reward per Episode"
            chart_rewards.y_axis.title = 'Reward'
            chart_rewards.x_axis.title = 'Episode'

            # Data columns: (2..3) => Avg Reward, Episode Reward
            data_rewards = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=max_row)
            # Category axis: Episode (column 1, rows 2..max_row)
            categories = Reference(ws, min_col=1, min_row=2, max_row=max_row)

            chart_rewards.add_data(data_rewards, titles_from_data=True)
            chart_rewards.set_categories(categories)
            chart_rewards.width = 20
            chart_rewards.height = 10

            # Place chart to the right of the data table
            ws.add_chart(chart_rewards, f"{get_column_letter(max_col+2)}2")  # G

            # --------------------------
            # Chart 2: Delay per Episode
            # --------------------------
            chart_delays = LineChart()
            chart_delays.title = "Delay per Episode"
            chart_delays.y_axis.title = 'Delay'
            chart_delays.x_axis.title = 'Episode'

            # Data columns: (4..5) => Avg Delay, Episode Delay
            data_delays = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=max_row)

            chart_delays.add_data(data_delays, titles_from_data=True)
            chart_delays.set_categories(categories)
            chart_delays.width = 20
            chart_delays.height = 10

            # Place second chart below the first one
            ws.add_chart(chart_delays, f"{get_column_letter(max_col+2)}20")


        elif sheet_name == 'Summary':
            # Column indices (1-based) for existing summary columns
            failure_col_idx = 5  # E = FailureRate
            episode_col_idx = 1  # A
            avg_col_idx = 6      # F AVG_Failure

            # Add headers for computed rolling average columns
            ws.cell(row=1, column=avg_col_idx, value='normal_AVG_Failure')
            ws.cell(row=1, column=avg_col_idx+1, value='AVG_Failure')

            # Read FailureRate values from the sheet (starting at row 2)
            failure_values = [row[0] for row in ws.iter_rows(min_row=2, min_col=failure_col_idx, max_col=failure_col_idx, values_only=True)]

            # Compute rolling average (window size 40, i.e., last 40 points)
            avg_values = []
            for i in range(len(failure_values)):
                if i >= 39:
                    avg = np.mean(failure_values[i-39:i+1])
                else:
                    avg = np.mean(failure_values[:i+1])
                avg_values.append(avg)

            # Write computed rolling averages back into the sheet
            for i, val in enumerate(avg_values, start=2):
                ws.cell(row=i, column=avg_col_idx, value=val)
                # Compute AVG_Failure as scaled by column 4 (likely "Total" or similar) / 100
                ws.cell(row=i, column=avg_col_idx+1, value=val * ws.cell(row=i, column=4).value/100)  # calculating AVG_Failure

            # -------------------------------
            # Line chart: Rolling Failure Avg
            # -------------------------------
            line_chart = LineChart()
            line_chart.title = "normal AVG Failure Rate Over Episodes"
            line_chart.style = 6
            line_chart.x_axis.title = 'Episode'
            line_chart.y_axis.title = 'normal_AVG_Failure'

            line_chart.width = 20
            line_chart.height = 10
            line_chart.legend = None

            # X-axis: episodes from column A
            episodes = Reference(ws, min_col=episode_col_idx, min_row=2, max_row=len(failure_values)+1)
            # Y-axis: rolling average failure from column F (including header row 1)
            avg_failure_ref = Reference(ws, min_col=avg_col_idx, min_row=1, max_row=len(failure_values)+1)

            line_chart.add_data(avg_failure_ref, titles_from_data=True)
            line_chart.set_categories(episodes)

            # Place chart to the right of the table
            ws.add_chart(line_chart, f"{get_column_letter(ws.max_column+2)}2")


    # Determine the directory of the current script file
    current_dir = os.path.dirname(os.path.abspath(__file__))

    # NEW: Determine scenario tag used in results directory naming
    scenario = getattr(params, "scenario", "base")

    # Determine probability parameter p based on scenario type
    if scenario == "trajectory_noise":
        p = getattr(params, "trajectory_noise_p", 0.0)
    elif scenario == "missing_data":
        p = getattr(params, "missing_data_p", 0.0)
    elif scenario == "base":
        p = 0.0
    else:
        # Default to zero if scenario is unknown
        p = 0.0

    # Build model tag used in folder naming (replace '.' to keep folder name clean)
    model_tag = f"{params.model_summary}_{p:.2f}".replace(".", "_")

    # Create results directory path: <SCENARIO_TYPE>_results/<scenario>/<model_tag>
    results_dir = os.path.join(current_dir, f"{params.SCENARIO_TYPE}_results", scenario,model_tag)
    os.makedirs(results_dir, exist_ok=True)


    # Load server info from an Excel file depending on homogeneous/heterogeneous scenario type
    fname = os.path.join(current_dir, 'homogeneous_server_info.xlsx') if params.SCENARIO_TYPE == 'homogeneous' else os.path.join(current_dir, 'heterogeneous_server_info.xlsx')
    sheet_name = f'{params.SCENARIO_TYPE.capitalize()}_Permutation_{params.Permutation_Number}'
    server_info = pd.read_excel(fname, sheet_name=sheet_name)

    # Load task parameters and server info into DataFrames
    task_df = pd.read_excel(os.path.join(current_dir, 'task_parameters.xlsx'))
    df_tasks = pd.DataFrame(task_df)
    df_servers = pd.DataFrame(server_info)

    
    # Define the global (aggregated) output Excel filename
    global_filename = os.path.join(results_dir, f"Global_Permutation_{params.Permutation_Number}_{params.model_summary}.xlsx")

    # If global file exists, read existing episodes from Logs sheet to prevent duplicates
    if os.path.exists(global_filename):
        existing_global = pd.read_excel(global_filename, sheet_name='Logs')
        existing_episodes = set(existing_global['Episode'].tolist())
    else:
        existing_episodes = set()

    # Keep only logs/assignments whose episode is not already present in the existing file
    new_logs = [log for log in log_data_global if log[0] not in existing_episodes]
    new_assignments = [a for a in task_Assignments_info_global if a[0] not in existing_episodes]

    # Convert new global logs to a DataFrame with fixed column mapping
    df_new_logs = pd.DataFrame([{'Episode': log[0], 'Avg Reward': log[1], 'Episode Reward': log[2], 'Avg Delay': log[3],'Episode Delay': log[4]} for log in new_logs])

    # Convert new assignment records to DataFrame with specified columns
    df_new_assignments = pd.DataFrame(new_assignments, columns=[ 
        'episode', 'task_id', 'vehicle_id', 'original_rsu', 'submitted_time',
        'selected_RSU', 'start_executing', 'final_rsu', 'finished_time', 'executaion_status', 'deadline_flag', 'final_status'
    ])

    # Build summary table: count Success/Failure per episode based on final_status
    summary_df = df_new_assignments.groupby(['episode', 'final_status']).size().unstack(fill_value=0).rename(columns={'f': 'Failure', 's': 'Success'}).reset_index()
    summary_df = summary_df.reindex(columns=['episode', 'Failure', 'Success'], fill_value=0)
    summary_df['Total'] = summary_df['Success'] + summary_df['Failure']
    summary_df['FailureRate'] = 100 * summary_df['Failure'] / summary_df['Total'].replace(0, 1)
    summary_df = summary_df.fillna(0) 

    if not os.path.exists(global_filename):
        # Create Params DataFrame from params object's attributes
        df_params = pd.DataFrame(list(vars(params).items()), columns=['Parameter', 'Value'])

        # Only keep the latest episode in TaskAssignments on initial creation
        ep_col = _detect_episode_col(df_new_assignments)
        last_ep = df_new_assignments[ep_col].max()
        df_last_assignments = df_new_assignments[df_new_assignments[ep_col] == last_ep].copy()

        # Write all sheets (Params, Tasks, Servers, Logs, TaskAssignments, Summary) into a new Excel file
        with pd.ExcelWriter(global_filename) as writer:
            df_params.to_excel(writer, sheet_name='Params', index=False)
            df_tasks.to_excel(writer, sheet_name='Tasks', index=False)
            df_servers.to_excel(writer, sheet_name='Servers', index=False)
            df_new_logs.to_excel(writer, sheet_name='Logs', index=False)
            df_last_assignments.to_excel(writer, sheet_name='TaskAssignments', index=False)  # ← only latest episode
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

    else:
        # Append new logs to Logs sheet
        append_df_to_sheet(global_filename, df_new_logs, 'Logs')

        # Do NOT append TaskAssignments; instead replace it with latest-episode-only records
        #append_df_to_sheet(global_filename, df_new_assignments, 'TaskAssignments')
        write_task_assignments_latest(global_filename, df_new_assignments, base_sheet_name='TaskAssignments')

        # Append summary rows to Summary sheet
        append_df_to_sheet(global_filename, summary_df, 'Summary')

    # Reload workbook to add/refresh charts on Logs and Summary sheets
    wb = load_workbook(global_filename)
    if 'Logs' in wb.sheetnames:
        add_charts_to_sheet(wb, 'Logs')
    if 'Summary' in wb.sheetnames:
        add_charts_to_sheet(wb, 'Summary')
    wb.save(global_filename)
    print("Global log updated.")

    # Identify cloud servers (used for RSU-level server sheet aggregation)
    cloud_servers = df_servers[df_servers['Server_Type'] == 'Cloud']

    # Iterate through each RSU and write/update its own Excel file
    for rsu_id in rsu_logs_dict.keys():
        filename_rsu = os.path.join(results_dir, f"{rsu_id}_Permutation_{params.Permutation_Number}_{params.model_summary}.xlsx")

        # Fetch RSU-local logs and assignments from dictionaries
        local_logs = rsu_logs_dict[rsu_id]
        local_assignments = rsu_assignments_dict[rsu_id]

        # If RSU file exists, detect existing episodes to avoid duplication
        if os.path.exists(filename_rsu):
            existing_rsu = pd.read_excel(filename_rsu, sheet_name='Logs')
            existing_episodes_rsu = set(existing_rsu['Episode'].tolist())
        else:
            existing_episodes_rsu = set()

        # Filter only new RSU logs/assignments for episodes not already present
        new_logs_rsu = [log for log in local_logs if log[0] not in existing_episodes_rsu]
        new_assignments_rsu = [a for a in local_assignments if a[0] not in existing_episodes_rsu]

        # Convert RSU logs and assignments to DataFrames
        df_logs_rsu = pd.DataFrame([{'Episode': log[0], 'Avg Reward': log[1], 'Episode Reward': log[2], 'Avg Delay': log[3],'Episode Delay': log[4]} for log in new_logs_rsu])
        df_assignments_rsu = pd.DataFrame(new_assignments_rsu, columns=[
            'episode', 'task_id', 'vehicle_id', 'Primary', 'Primary_Start', 'Primary_End', 'Primary_Status',
            'Backup', 'Backup_Start', 'Backup_End', 'Backup_Status', 'Z', 'executaion_status'
        ])

        # RSU servers = servers belonging to this RSU + all cloud servers
        rsu_servers = pd.concat([df_servers[df_servers['RSU_ID'] == rsu_id], cloud_servers], ignore_index=True)

        # Build RSU summary: count Success/Failure per episode based on executaion_status
        summary_rsu = df_assignments_rsu.groupby(['episode', 'executaion_status']).size().unstack(fill_value=0).rename(columns={'f': 'Failure', 's': 'Success'}).reset_index()
        summary_rsu = summary_rsu.reindex(columns=['episode', 'Failure', 'Success'], fill_value=0)
        summary_rsu['Total'] = summary_rsu['Success'] + summary_rsu['Failure']
        summary_rsu['FailureRate'] = 100 * summary_rsu['Failure'] / summary_rsu['Total'].replace(0, 1)
        summary_rsu = summary_rsu.fillna(0)       # For RSU sheets

        if not os.path.exists(filename_rsu):
            # Only keep the latest episode in 'TaskAssignments' on initial creation
            ep_col_rsu = _detect_episode_col(df_assignments_rsu) if len(df_assignments_rsu) else 'episode'
            last_ep_rsu = df_assignments_rsu[ep_col_rsu].max() if len(df_assignments_rsu) else None
            df_last_assignments_rsu = (
                df_assignments_rsu[df_assignments_rsu[ep_col_rsu] == last_ep_rsu].copy()
                if last_ep_rsu is not None else df_assignments_rsu.copy()
            )

            # Create a new RSU Excel file with Servers, Logs, TaskAssignments, Summary sheets
            with pd.ExcelWriter(filename_rsu) as writer:
                rsu_servers.to_excel(writer, sheet_name='Servers', index=False)
                df_logs_rsu.to_excel(writer, sheet_name='Logs', index=False)
                df_last_assignments_rsu.to_excel(writer, sheet_name='TaskAssignments', index=False)
                summary_rsu.to_excel(writer, sheet_name='Summary', index=False)
        else:
            # Append new RSU logs
            append_df_to_sheet(filename_rsu, df_logs_rsu, 'Logs')

            # Replace RSU TaskAssignments sheet with latest-episode-only records
            write_task_assignments_latest(filename_rsu, df_assignments_rsu, base_sheet_name='TaskAssignments')

            # Append RSU summary rows
            append_df_to_sheet(filename_rsu, summary_rsu, 'Summary')

        # Reload RSU workbook to add/refresh charts
        wb_rsu = load_workbook(filename_rsu)
        if 'Logs' in wb_rsu.sheetnames:
            add_charts_to_sheet(wb_rsu, 'Logs')
        if 'Summary' in wb_rsu.sheetnames:
            add_charts_to_sheet(wb_rsu, 'Summary')
        wb_rsu.save(filename_rsu)
        print(f"{rsu_id} log updated.")
